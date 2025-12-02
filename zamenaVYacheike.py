import os
import openpyxl

# Путь к файлу Замены.xlsx
replacement_file_path = r'C:\Users\Михаил\Desktop\Артикулы\Замены.xlsx'
# Путь к папке с файлами для замены
target_folder_path = r'C:\Users\Михаил\Desktop\Артикулы\Заливка2'

# Открываем файл Замены.xlsx
wb_replacements = openpyxl.load_workbook(replacement_file_path)
ws_replacements = wb_replacements.active

# Проходим по всем строкам в файле Замены.xlsx
print("Начал")
krug=0
for row in ws_replacements.iter_rows(min_row=1, max_col=2, values_only=True):
    krug+=1
    print(krug)
    search_value = row[0]  # Значение для поиска из первого столбца
    replace_value = row[1]  # Значение для замены из второго столбца

    # Если значение для поиска пустое, выходим из цикла
    if search_value is None:
        break

    # Проходим по всем файлам в папке Заливка
    for filename in os.listdir(target_folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(target_folder_path, filename)
            wb_target = openpyxl.load_workbook(file_path)
            ws_target = wb_target.active

            # Проходим по всем ячейкам в целевом файле
            for row in ws_target.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and search_value in cell.value:
                        # Заменяем часть содержимого ячейки
                        cell.value = cell.value.replace(search_value, replace_value)

            # Сохраняем изменения в целевом файле
            wb_target.save(file_path)
            wb_target.close()

# Закрываем файл Замены.xlsx
wb_replacements.close()
print("Выполнено")
