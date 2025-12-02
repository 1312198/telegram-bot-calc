import os
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# Путь к папке
folder_path = r'C:\Users\Михаил\Desktop\Папка Z'

# Поиск Excel файла в папке
def find_excel_file(folder):
    for file in os.listdir(folder):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            return os.path.join(folder, file)
    return None

# Основной процесс обработки
def process_excel(file_path):
    try:
        # Открываем файл
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Удаление пустых строк в столбце B (2)
        rows_to_delete = []
        ip = 0
        for row in range(2, 137305):  # диапазон от 2 до 137304
            if ws.cell(row=row, column=2).value is None:
                ip += 1
                rows_to_delete.append(row)
                print (f"Для удаления {ip}.")
        
        # Удаляем строки в обратном порядке
        yd = 0
        for row in sorted(rows_to_delete, reverse=True):
            yd += 1
            ws.delete_rows(row)
            print (f"Удалено. {yd}")
        
        # Обработка столбца B (2) для переноса доменов
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=2).value
            if cell_value:
                if '.ru' in str(cell_value) or '.com' in str(cell_value) or '.рф' in str(cell_value):
                    # Переносим значение в столбец A (1)
                    ws.cell(row=row, column=1).value = cell_value
                    ws.cell(row=row, column=2).value = None
                    print ("Перенес.")
        
        # Сохраняем изменения
        wb.save(file_path)
        print("Файл успешно обработан и сохранен")
    
    except FileNotFoundError:
        print("Файл не найден")
    except InvalidFileException:
        print("Ошибка при открытии файла Excel")
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")
    finally:
        # Закрываем рабочую книгу
        try:
            wb.close()
            print("Файл закрыт")
        except:
            pass

# Основной запуск
excel_file = find_excel_file(folder_path)
if excel_file:
    print(f"Найден файл: {excel_file}")
    process_excel(excel_file)
else:
    print("Excel файл не найден в указанной папке")