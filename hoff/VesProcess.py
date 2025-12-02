import os
import requests
from openpyxl import load_workbook
from PIL import Image

# 1. Запрашиваем адрес к нужной нам папке.
folder_path = input("Введите путь к папке (исходной): ")

# 2. В этой папке создаем папку - "Фото".
photos_folder = os.path.join(folder_path, "Фото")
os.makedirs(photos_folder, exist_ok=True)

# 3. По полученному в п.1 адресу находим Excel файл.
excel_file = None
for file in os.listdir(folder_path):
    if file.endswith('.xlsx') and not file.startswith('~$'):
        excel_file = os.path.join(folder_path, file)
        break

if excel_file is None:
    print("Excel файл не найден.")
    exit()

# 4. В файле из первой ячейки первого столбца забираем данные.
workbook = load_workbook(excel_file)
sheet = workbook.active

# 5-7. Проходим по строкам и столбцам
row = 1
while True:
    title_cell = sheet.cell(row=row, column=1)
    if title_cell.value is None:
        break  # Если пустая ячейка в столбце 1, выходим из цикла.

    title = title_cell.value
    col = 2
    while True:
        link_cell = sheet.cell(row=row, column=col)
        if link_cell.value is None:
            break  # Если пустая ячейка в текущем столбце, выходим из цикла.

        # 5. Скачиваем фото по ссылке.
        img_url = link_cell.value
        img_name = f"{title}_{col - 1}.jpg"
        img_path = os.path.join(photos_folder, img_name)

        try:
            response = requests.get(img_url)
            response.raise_for_status()  # Проверяем на ошибки
            with open(img_path, 'wb') as img_file:
                img_file.write(response.content)
            print(f"Скачано: {img_name}")
        except Exception as e:
            print(f"Ошибка при скачивании {img_name}: {e}")

        col += 1  # Переходим к следующему столбцу.

    row += 1  # Переходим к следующей строке.

# 8. Закрываем файл.
workbook.close()
print("Работа по скачиванию фото окончена.")

# 9. Обработка изображений.
for filename in os.listdir(photos_folder):
    if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
        img_path = os.path.join(photos_folder, filename)
        image = Image.open(img_path)

        # 10. Проверяем размеры изображения.
        width, height = image.size
        if width < 1000 and height < 1000:
            scale_factor = 1000 / min(width, height)
            image = image.resize((int(width * scale_factor), int(height * scale_factor)))

        # Сохраняем результат.
        image.save(img_path, image.format)

# 11. Обработка изображений с учетом фона 3:2.
for filename in os.listdir(photos_folder):
    if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
        img_path = os.path.join(photos_folder, filename)
        image = Image.open(img_path)

        width, height = image.size
        if height >= width:
            new_width = int(height * 1.5)
            background = Image.new('RGB', (new_width, height), color=(255, 255, 255))
            box = (int((new_width - width) / 2), 0)
            background.paste(image, box)
        else:
            new_height = int(width * 2 / 3)
            background = Image.new('RGB', (width, new_height), color=(255, 255, 255))
            box = (0, int((new_height - height) / 2))
            background.paste(image, box)

        # Сохраняем результат.
        background.save(img_path, image.format)

# 12. Заканчиваем работу программы.
print("Все готово.")
